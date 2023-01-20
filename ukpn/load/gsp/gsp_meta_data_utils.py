"""This class is ued to retrieve data through API calls"""
import json
import logging
import os
from pathlib import Path
from pprint import pprint
from typing import Optional, Union

import numpy as np
import pandas as pd
import requests
import xarray as xr
from openpyxl import load_workbook

from ukpn.scripts.resample_data import interpolation_pandas, load_csv_to_pandas

logger = logging.getLogger(__name__)


def get_metadata_from_ukpn_api(
    api_url: str,
    eastings: Optional[str] = None,
    northings: Optional[str] = None,
    print_data: bool = False,
):
    """
    This function retrievs metadata through api calls

    Args:
        api_url: The api url link that emiits json format data
        print_data: Optional to choose printing the data
        eastings: eastings value of the pv solar farm
        northings: Northings value of the pv solar farm
    """

    response_api = requests.get(api_url)
    while True:
        if response_api == 200:
            logger.info(f"The api resposne {response_api} is successful")
        else:
            logger.warning(f"The api resposne {response_api} is unsuccessul")
            logger.info(f"Please enter the correct {'url'}")
            break

    # Get the data from the resposne
    raw_data = response_api.text

    # Parse the data into json format
    data_json = json.loads(raw_data)

    # Getting all the records
    data_records = data_json["records"]
    first_record = data_json["records"][0]

    if print_data:
        pprint(first_record)

    pv_site_dict_index = []
    # From the list of dictionaries
    for i in range(len(data_records)):
        if isinstance(data_records[i], dict):
            fields = data_records[i]["fields"]
            if isinstance(fields, dict):
                if (
                    fields["location_x_coordinate_eastings_where_data_is_held"] == eastings
                    and fields["location_y_coordinate_northings_where_data_is_held"] == northings
                ) is True:
                    pv_site_dict_index.append(i)

    # CHecking if there are any sites matching the coordinates
    if len(pv_site_dict_index) == 0:
        logger.info(f"There are no PV sites matching with eastinngs: {eastings}")
        logger.info(f"There are no PV sites matching with northings: {northings}")
        return None
    else:
        # Getting the required data from Eastings and Northings
        data_json = data_records[pv_site_dict_index[0]]

        return data_json


def get_metadata_from_ukpn_xlsx(
    link_of_ecr_excel: str,
    local_path: Path[Union, str],
    eastings: Optional[str] = None,
    northings: Optional[str] = None,
) -> pd.DataFrame:
    """Download and load the ECR file from the link provided below

    For direct download, opne this link-
    https://media.umbraco.io/uk-power-networks/0dqjxaho/embedded-capacity-register.xlsx

    Args:
        link_of_ecr_excel: Link shown above
        local_path: The folder where the file needs to get downloaded
        eastings: eastings value of the pv solar farm
        northings: Northings value of the pv solar farm
    """
    # Download and store the excel sheet in a location
    resp = requests.get(link_of_ecr_excel)
    local_path = os.path.join(local_path, "ecr.xlsx")
    with open(local_path, "wb") as output:
        output.write(resp.content)

    # Read the excel sheet
    wb = load_workbook(local_path, read_only=True, keep_links=False)

    # The sheet need and its name according to UKPN is "Register Part 1"
    file_name = "Register Part 1"
    for text in wb.sheetnames:
        if file_name in text:
            df = pd.read_excel(local_path, sheet_name=text, skiprows=1)

    # Filtering the data based on the eastings and northings provided
    for text in df.columns:
        if "Eastings" in text:
            df = df[df[text] == np.float64(eastings)].reset_index()

    return df


def metadata_df_to_netcdf(
    path_to_ukpn_timeseries: str, input_dataframe: Optional[pd.DataFrame] = None
) -> xr.Dataset:

    # Loading the UKPN time series data into a dataframe
    original_df = load_csv_to_pandas(path_to_file=path_to_ukpn_timeseries)

    # Interpolating time series
    interpolated_df = interpolation_pandas(original_df=original_df)

    # Getting all the time series dates
    interpolated_timeseries = interpolated_df.index.values
    interpolated_data_values = interpolated_df[interpolated_df.columns[0]].values

    # Creating an Xarray data array
    final_xarray = xr.DataArray(
        data=interpolated_data_values,
        dims="time_utc",
        coords={"time_utc": interpolated_timeseries},
        attrs={"Description": " This Data array consists of time-series data from UKPN"},
    )

    # Getiing the column names
    # TODO
    required_column_names = ["Eastings", "Northings", "Maximum Export Capacity"]

    return final_xarray


def construct_url(
    dataset_name: str = "embedded-capacity-register",
    list_of_facets=None,
    refiners=None,
    refine_values=None,
):
    """This function constructs a downloadble url of JSON data

    For more information, please visit
    - https://ukpowernetworks.opendatasoft.com/pages/home/

    Args:
        dataset_name: Name of the dataset that needs to be downloaded, defined by UKPN
        list_of_facets: List of facets that needs to be included in the JSON data
        refiners: list of refiner terms that needs to refined from the JSON data
        refine_values: List of refine values of the refiners

    Note:
        refiners and refine values needs to be exactly mapped
    """
    # Constructing a base url
    base_url = "https://ukpowernetworks.opendatasoft.com/api/records/1.0/search/?dataset="
    base_url = base_url + dataset_name

    # A seperator in the url
    seperator = "&"

    # A questionare in the url
    questionare = "q="

    # A facet questionare in the url
    facet_questionare = "facet="

    # Constructing a facet string from the list of facets
    facet_str = [facet_questionare + x for x in list_of_facets]
    facet_str = seperator.join(facet_str)
    facet_str = str(questionare + seperator + facet_str)

    # Constructing a refiner string to refine the JSON data
    refine_questionare = "refine."
    refiners = [refine_questionare + x for x in refiners]
    refiners = list(map(lambda x, y: x + str("=") + y, refiners, refine_values))
    refiners = seperator.join(refiners)

    # Constructing the final url
    final_url = [base_url, facet_str, refiners]
    final_url = seperator.join(final_url)
    return final_url

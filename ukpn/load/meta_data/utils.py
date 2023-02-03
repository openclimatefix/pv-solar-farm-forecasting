"""This class is ued to retrieve data through API calls"""
import json
import logging
import os
from glob import glob
from pathlib import Path
from pprint import pprint
from typing import List, Optional, Union, Dict

import numpy as np
import pandas as pd
import requests
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def get_metadata_from_ukpn_api(
    api_url: str
):
    """
    This function retrievs metadata through api calls

    Args:
        api_url: The api url link that emiits json format data
    """

    response_api = requests.head(api_url)
    if not response_api.status_code == 200:
        logger.info(f"The response from the link {api_url} is unsuccessful")
        return None
    else:
        logger.info(f"The resposne from the link {api_url} is successful")

        # Get the data from the resposne
        response_api = requests.get(api_url)
        raw_data = response_api.text

        # Parse the data into json format
        data_json = json.loads(raw_data)
        return data_json


def get_metadata_from_ukpn_xlsx(
    link_of_ecr_excel: str,
    local_path: Path[Union, str],
    eastings: Optional[str] = None,
    northings: Optional[str] = None,
) -> pd.DataFrame:
    """Download and load the ECR file from the link provided below

    For direct download, opne this link-
    https://media.umbraco.io/uk-power-networks/0dqjxaho/embedded-capacity-register.xlsx

    Args:
        link_of_ecr_excel: Link shown above
        local_path: The folder where the file needs to get downloaded
        eastings: eastings value of the pv solar farm
        northings: Northings value of the pv solar farm
    """
    # Download and store the excel sheet in a location
    resp = requests.get(link_of_ecr_excel)
    local_path = os.path.join(local_path, "ecr.xlsx")
    with open(local_path, "wb") as output:
        output.write(resp.content)

    # Read the excel sheet
    wb = load_workbook(local_path, read_only=True, keep_links=False)

    # The sheet need and its name according to UKPN is "Register Part 1"
    file_name = "Register Part 1"
    for text in wb.sheetnames:
        if file_name in text:
            df = pd.read_excel(local_path, sheet_name=text, skiprows=1)

    # Filtering the data based on the eastings and northings provided
    for text in df.columns:
        if "Eastings" in text:
            df = df[df[text] == np.float64(eastings)].reset_index()

    return df


def construct_url(
    dataset_name: str = "embedded-capacity-register",
    list_of_facets: List = [
        "grid_supply_point",
        "licence_area",
        "energy_conversion_technology_1",
        "flexible_connection_yes_no",
        "connection_status",
        "primary_resource_type_group",
    ],
    refine_facet:List = None,
    refine_facet_values:List = None,
    number_of_records: str = '5000',
):
    """This function constructs a downloadable url of UKPN PV JSON data

    For more information, please visit
    - https://ukpowernetworks.opendatasoft.com/pages/home/

    Args:
        dataset_name: Name of the dataset that needs to be downloaded, defined by UKPN
        list_of_facets: LIst of facets that needs to be included in the JSON data
        refine_facet: List of refiners that needs to be included in the JSON data
        refine_facet_values: List of refine values of the refiners
        number_of_records: Number of records needed to be extracted

    Note:
        refiners and refine values needs to be exactly mapped
    """

    # Constructing a base url
    base_url = "https://ukpowernetworks.opendatasoft.com/api/records/1.0/search/?dataset="
    base_url = base_url + dataset_name

    # A seperator in the url
    seperator = "&"

    # A questionare in the url
    questionare = "q="

    # Number of records needed to be extracted
    total_rows = str("&rows="+number_of_records)

    # A facet questionare in the url
    facet_questionare = "facet="

    # Constructing a facet string from the list of facets
    facet_str = [facet_questionare + x for x in list_of_facets]
    facet_str = seperator.join(facet_str)
    facet_str = str(questionare + total_rows + seperator + facet_str)
    
    # Constructing a refiner string to refine the JSON data
    refiners = ["energy_conversion_technology_1"]
    refine_values = ["Photovoltaic"]

    if refine_facet is None:
        pass
    else:
        # Sanity checs
        sanity_check = len(refine_facet) == len(refine_facet_values)
        if not sanity_check:
            logger.debug(f"The URL is invalid, total {refiners} should be equal to {refine_values}")
            return None             
        [refiners.append(x) for x in refine_facet]
        [refine_values.append(x) for x in refine_facet_values]

    refine_questionare = "refine."
    refiners = [refine_questionare + x for x in refiners]
    refiners = list(map(lambda x, y: x + str("=") + y, refiners, refine_values))
    refiners = seperator.join(refiners)

    # Getting the entire UKPN records
    final_url = [base_url, facet_str, refiners]
    final_url = seperator.join(final_url)

    return final_url

def get_gsp_names(
    folder_destination: str,
    file_format: str = "*.csv"
    ) -> Dict:
    """
    This methods takes all the csv file names and gives their
    respective GSP names from the UKPN dashboard
    
    Args:
        folder_destination: The destination folder that has all the csv files
        file+fomat: The file format that needs to be searched.
    """
    # Getting all the records
    api_url = construct_url()
    data_json = get_metadata_from_ukpn_api(api_url = api_url)
    # Getting all the gsp_names
    if isinstance(data_json, Dict):
        gsp_facet_group = data_json["facet_groups"]
    else:
        return None
    
    # Getting the Grid supply point group
    for group in gsp_facet_group:
        if group["name"] == "grid_supply_point":
            gsp_group = group["facets"]
    # Getting the GSP names
    gsp_names = []
    for each_gsp in gsp_group:
        gsp_names.append(each_gsp["name"])

    # Getting the file paths
    file_paths = os.path.join(folder_destination, file_format)
    file_paths = [x for x in glob(file_paths)]

    # Getting each file name
    gsp_name_dict = {}
    for file_path in file_paths:
        base_name = os.path.basename(file_path)
        file_name = os.path.splitext(base_name)[0]
        i = 0
        while i < len(gsp_names):
            if file_name.upper() in gsp_names[i]:
                # Joining with + seperator
                gsp_name = gsp_names[i].split(' ')
                gsp_name = '+'.join(gsp_name)
                gsp_name_dict[file_path] = gsp_name
            i+=1
    
    return gsp_name_dict
